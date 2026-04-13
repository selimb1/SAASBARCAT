"""
Exportador para Holistor.
Genera archivo Excel (.xlsx) con dos hojas:
  - COMPROBANTES_COMPRAS: para el módulo de Compras / Cuentas a Pagar
  - ASIENTOS: entradas de asiento para el módulo de Contabilidad (opcional)
"""
import io
from datetime import date
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.models.models import Comprobante

TIPO_COMP_NOMBRES = {
    "A": "Factura A",
    "B": "Factura B",
    "C": "Factura C",
    "M": "Factura M",
    "E": "Factura E",
    "FCE_A": "Factura Crédito Electrónica A",
    "FCE_B": "Factura Crédito Electrónica B",
    "ND_A": "Nota de Débito A",
    "ND_B": "Nota de Débito B",
    "ND_C": "Nota de Débito C",
    "NC_A": "Nota de Crédito A",
    "NC_B": "Nota de Crédito B",
    "NC_C": "Nota de Crédito C",
    "RECIBO": "Recibo",
    "TICKET": "Tique",
    "REMITO": "Remito",
}


def _nombre_tipo(comp: Comprobante) -> str:
    tipo = comp.tipo_comprobante or ""
    return TIPO_COMP_NOMBRES.get(tipo, f"Factura {tipo}" if tipo else "Desconocido")


def _num_comprobante(comp: Comprobante) -> str:
    pv = (comp.punto_venta or "0001").zfill(4)
    nro = (comp.numero_comprobante or "0").zfill(8)
    return f"{pv}-{nro}"


def _fmt_fecha(d: Optional[date]) -> Optional[str]:
    if not d:
        return None
    return d.strftime("%d/%m/%Y")


def _v(valor) -> float:
    return float(valor or 0)


def _style_header(ws, row: int, cols: int, title: str = ""):
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        bottom=Side(style="medium", color="4A90D9"),
    )
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def generar_holistor(
    comprobantes: list[Comprobante],
    cuenta_gasto_defecto: str = "5.1.01.001",
    empresa_codigo: str = "001",
    incluir_asientos: bool = True,
) -> bytes:
    """
    Genera archivo Excel para Holistor con las hojas de Comprobantes y Asientos.
    """
    wb = Workbook()

    # ── Hoja 1: COMPROBANTES_COMPRAS ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "COMPROBANTES_COMPRAS"

    encabezados = [
        "TIPO_COMP", "NUM_COMP", "FECHA", "CUIT_PROV", "PROV_NOMBRE",
        "NETO_21", "IVA_21", "NETO_105", "IVA_105",
        "EXENTO", "NO_GRAV", "PERC_IVA", "PERC_IIBB", "RETENC",
        "TOTAL", "CAE", "COND_PAGO", "CUENTA_GASTO", "OBSERV"
    ]

    # Título
    ws1.merge_cells(f"A1:{get_column_letter(len(encabezados))}1")
    titulo_cell = ws1["A1"]
    titulo_cell.value = "contabilizAR — Importación Holistor — Comprobantes de Compra"
    titulo_cell.font = Font(bold=True, color="1E3A5F", size=12)
    titulo_cell.alignment = Alignment(horizontal="center")
    titulo_cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    ws1.row_dimensions[1].height = 22

    # Encabezados
    for col_idx, header in enumerate(encabezados, 1):
        ws1.cell(row=2, column=col_idx, value=header)
    _style_header(ws1, 2, len(encabezados))
    ws1.row_dimensions[2].height = 18

    # Anchos de columna
    anchos = [18, 16, 12, 18, 35, 14, 12, 14, 12, 12, 12, 12, 12, 12, 14, 18, 16, 18, 25]
    for i, ancho in enumerate(anchos, 1):
        ws1.column_dimensions[get_column_letter(i)].width = ancho

    # Datos
    alt_fill = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
    num_format = '#,##0.00'
    date_format = 'DD/MM/YYYY'

    for row_idx, comp in enumerate(comprobantes, 3):
        fill = alt_fill if row_idx % 2 == 0 else None

        perc_iva = 0.0
        perc_iibb = 0.0
        if comp.percepciones_detalle:
            for p in comp.percepciones_detalle:
                if p.get("tipo") == "IVA":
                    perc_iva = float(p.get("importe", 0))
                elif p.get("tipo") == "IIBB":
                    perc_iibb = float(p.get("importe", 0))
        elif _v(comp.importe_percepciones) > 0:
            perc_iibb = _v(comp.importe_percepciones)

        cond_pago = {
            "CONTADO": "Contado",
            "CUENTA_CORRIENTE": "Cuenta Corriente",
            "CREDITO": "Crédito",
        }.get(comp.condicion_venta or "CONTADO", "Contado")

        fila = [
            _nombre_tipo(comp),
            _num_comprobante(comp),
            comp.fecha_emision,
            comp.cuit_emisor or "",
            (comp.razon_social_emisor or "")[:60],
            _v(comp.importe_neto_gravado),
            _v(comp.iva_21),
            0.0,  # neto 10.5 (si el usuario tiene desglose separado)
            _v(comp.iva_105),
            _v(comp.importe_exento),
            _v(comp.importe_no_gravado),
            perc_iva,
            perc_iibb,
            _v(comp.importe_retenciones),
            _v(comp.importe_total),
            comp.cae or "",
            cond_pago,
            cuenta_gasto_defecto,
            comp.concepto_descripcion or "",
        ]

        for col_idx, valor in enumerate(fila, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=valor)
            if fill:
                cell.fill = fill
            if isinstance(valor, float):
                cell.number_format = num_format
                cell.alignment = Alignment(horizontal="right")
            elif isinstance(valor, date):
                cell.number_format = date_format
                cell.alignment = Alignment(horizontal="center")

    # ── Hoja 2: ASIENTOS (opcional) ─────────────────────────────────────────
    if incluir_asientos:
        ws2 = wb.create_sheet(title="ASIENTOS")
        enc_asientos = ["FECHA", "COD_ASIENTO", "DESCRIPCION", "CUENTA", "DEBE", "HABER", "CENTRO_COSTO", "AUXILIAR"]

        ws2.merge_cells(f"A1:{get_column_letter(len(enc_asientos))}1")
        titulo_as = ws2["A1"]
        titulo_as.value = "contabilizAR — Asientos para Holistor"
        titulo_as.font = Font(bold=True, color="1E3A5F", size=12)
        titulo_as.alignment = Alignment(horizontal="center")
        titulo_as.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

        for col_idx, header in enumerate(enc_asientos, 1):
            ws2.cell(row=2, column=col_idx, value=header)
        _style_header(ws2, 2, len(enc_asientos))

        anchos_as = [12, 14, 40, 16, 14, 14, 16, 20]
        for i, ancho in enumerate(anchos_as, 1):
            ws2.column_dimensions[get_column_letter(i)].width = ancho

        row_as = 3
        for idx, comp in enumerate(comprobantes, 1):
            cod_asiento = str(idx).zfill(6)
            desc_base = f"{_nombre_tipo(comp)} {_num_comprobante(comp)}"
            cuit_aux = comp.cuit_emisor or ""
            total = _v(comp.importe_total)

            # Fila débito: cuenta de gasto / mercadería
            if _v(comp.importe_neto_gravado) > 0:
                ws2.cell(row=row_as, column=1, value=comp.fecha_emision)
                ws2.cell(row=row_as, column=2, value=cod_asiento)
                ws2.cell(row=row_as, column=3, value=f"COMPRA {desc_base}")
                ws2.cell(row=row_as, column=4, value=cuenta_gasto_defecto)
                ws2.cell(row=row_as, column=5, value=_v(comp.importe_neto_gravado)).number_format = num_format
                ws2.cell(row=row_as, column=6, value=0.0)
                ws2.cell(row=row_as, column=8, value=cuit_aux)
                row_as += 1

            # IVA Crédito Fiscal
            if _v(comp.iva_21) > 0:
                ws2.cell(row=row_as, column=1, value=comp.fecha_emision)
                ws2.cell(row=row_as, column=2, value=cod_asiento)
                ws2.cell(row=row_as, column=3, value=f"IVA CF 21% {desc_base}")
                ws2.cell(row=row_as, column=4, value="1.5.01.001")
                ws2.cell(row=row_as, column=5, value=_v(comp.iva_21)).number_format = num_format
                ws2.cell(row=row_as, column=6, value=0.0)
                row_as += 1

            if _v(comp.iva_105) > 0:
                ws2.cell(row=row_as, column=1, value=comp.fecha_emision)
                ws2.cell(row=row_as, column=2, value=cod_asiento)
                ws2.cell(row=row_as, column=3, value=f"IVA CF 10.5% {desc_base}")
                ws2.cell(row=row_as, column=4, value="1.5.01.002")
                ws2.cell(row=row_as, column=5, value=_v(comp.iva_105)).number_format = num_format
                ws2.cell(row=row_as, column=6, value=0.0)
                row_as += 1

            # Haber: Proveedores a Pagar
            ws2.cell(row=row_as, column=1, value=comp.fecha_emision)
            ws2.cell(row=row_as, column=2, value=cod_asiento)
            ws2.cell(row=row_as, column=3, value=f"PROVEEDORES {desc_base}")
            ws2.cell(row=row_as, column=4, value="2.1.01.001")
            ws2.cell(row=row_as, column=5, value=0.0)
            ws2.cell(row=row_as, column=6, value=total).number_format = num_format
            ws2.cell(row=row_as, column=8, value=cuit_aux)
            row_as += 1

    # ── Output ───────────────────────────────────────────────────────────────
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
